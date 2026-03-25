"""
╔══════════════════════════════════════════════════════════════╗
║         NIFTY 50 STOCK TREND ANALYZER                       ║
║         3-Month Technical Analysis Tool                      ║
║         Outputs: Charts + Excel Report                       ║
╚══════════════════════════════════════════════════════════════╝

Requirements:
    pip install yfinance pandas numpy matplotlib seaborn openpyxl ta

Usage:
    python nifty50_analyzer.py
    python nifty50_analyzer.py --period 6mo     # change period
    python nifty50_analyzer.py --stock RELIANCE  # single stock deep-dive
"""

import argparse
import os
import sys
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as mpatches
from matplotlib.colors import LinearSegmentedColormap
import matplotlib.ticker as mticker
import seaborn as sns

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# NIFTY 50 CONSTITUENTS
# ─────────────────────────────────────────────────────────────────────────────
NIFTY50 = {
    "ADANIENT":   ("Adani Enterprises",      "Commodities"),
    "ADANIPORTS": ("Adani Ports & SEZ",       "Infrastructure"),
    "APOLLOHOSP": ("Apollo Hospitals",        "Healthcare"),
    "ASIANPAINT": ("Asian Paints",            "Consumer"),
    "AXISBANK":   ("Axis Bank",               "Banking"),
    "BAJAJ-AUTO": ("Bajaj Auto",              "Auto"),
    "BAJAJFINSV": ("Bajaj Finserv",           "Financial Services"),
    "BAJFINANCE": ("Bajaj Finance",           "Financial Services"),
    "BEL":        ("Bharat Electronics",      "Defence"),
    "BPCL":       ("BPCL",                    "Energy"),
    "BHARTIARTL": ("Bharti Airtel",           "Telecom"),
    "BRITANNIA":  ("Britannia Industries",    "FMCG"),
    "CIPLA":      ("Cipla",                   "Pharma"),
    "COALINDIA":  ("Coal India",              "Commodities"),
    "DRREDDY":    ("Dr. Reddy's Labs",        "Pharma"),
    "EICHERMOT":  ("Eicher Motors",           "Auto"),
    "GRASIM":     ("Grasim Industries",       "Cement"),
    "HCLTECH":    ("HCL Technologies",        "IT"),
    "HDFCBANK":   ("HDFC Bank",               "Banking"),
    "HDFCLIFE":   ("HDFC Life Insurance",     "Insurance"),
    "HEROMOTOCO": ("Hero MotoCorp",           "Auto"),
    "HINDALCO":   ("Hindalco Industries",     "Metals"),
    "HINDUNILVR": ("Hindustan Unilever",      "FMCG"),
    "ICICIBANK":  ("ICICI Bank",              "Banking"),
    "INDUSINDBK": ("IndusInd Bank",           "Banking"),
    "INFY":       ("Infosys",                 "IT"),
    "ITC":        ("ITC",                     "FMCG"),
    "JIOFIN":     ("Jio Financial Services",  "Financial Services"),
    "JSWSTEEL":   ("JSW Steel",               "Metals"),
    "KOTAKBANK":  ("Kotak Mahindra Bank",     "Banking"),
    "LT":         ("Larsen & Toubro",         "Infrastructure"),
    "LTIM":       ("LTIMindtree",             "IT"),
    "M&M":        ("Mahindra & Mahindra",     "Auto"),
    "MARUTI":     ("Maruti Suzuki",           "Auto"),
    "NESTLEIND":  ("Nestle India",            "FMCG"),
    "NTPC":       ("NTPC",                    "Energy"),
    "ONGC":       ("ONGC",                    "Energy"),
    "POWERGRID":  ("Power Grid Corp",         "Energy"),
    "RELIANCE":   ("Reliance Industries",     "Energy"),
    "SBILIFE":    ("SBI Life Insurance",      "Insurance"),
    "SBIN":       ("State Bank of India",     "Banking"),
    "SHRIRAMFIN": ("Shriram Finance",         "Financial Services"),
    "SUNPHARMA":  ("Sun Pharmaceutical",      "Pharma"),
    "TATACONSUM": ("Tata Consumer Products",  "FMCG"),
    "TATAMOTORS": ("Tata Motors",             "Auto"),
    "TATASTEEL":  ("Tata Steel",              "Metals"),
    "TCS":        ("Tata Consultancy Svcs",   "IT"),
    "TECHM":      ("Tech Mahindra",           "IT"),
    "TITAN":      ("Titan Company",           "Consumer"),
    "ULTRACEMCO": ("UltraTech Cement",        "Cement"),
    "WIPRO":      ("Wipro",                   "IT"),
}

SECTORS = sorted(set(v[1] for v in NIFTY50.values()))

# ─────────────────────────────────────────────────────────────────────────────
# COLOUR PALETTE
# ─────────────────────────────────────────────────────────────────────────────
PALETTE = {
    "bg":        "#0D1117",
    "surface":   "#161B22",
    "card":      "#1C2128",
    "border":    "#30363D",
    "text":      "#E6EDF3",
    "muted":     "#8B949E",
    "accent":    "#58A6FF",
    "green":     "#3FB950",
    "red":       "#F85149",
    "amber":     "#D29922",
    "purple":    "#BC8CFF",
    "teal":      "#39D353",
    "orange":    "#DB6D28",
}

SECTOR_COLORS = {
    "IT":                "#58A6FF",
    "Banking":           "#3FB950",
    "Financial Services":"#BC8CFF",
    "FMCG":              "#F0883E",
    "Energy":            "#D29922",
    "Auto":              "#39D353",
    "Pharma":            "#56D364",
    "Metals":            "#79C0FF",
    "Infrastructure":    "#FF7B72",
    "Consumer":          "#FFA657",
    "Telecom":           "#D2A8FF",
    "Healthcare":        "#7EE787",
    "Commodities":       "#E3B341",
    "Cement":            "#A5D6FF",
    "Defence":           "#FF6E6E",
    "Insurance":         "#B3F0FF",
}


# ─────────────────────────────────────────────────────────────────────────────
# TECHNICAL INDICATORS (pure numpy/pandas — no ta library needed)
# ─────────────────────────────────────────────────────────────────────────────
def compute_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    delta = series.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_gain = gain.ewm(com=period - 1, min_periods=period).mean()
    avg_loss = loss.ewm(com=period - 1, min_periods=period).mean()
    rs  = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi


def compute_macd(series: pd.Series,
                 fast: int = 12, slow: int = 26, signal: int = 9):
    ema_fast   = series.ewm(span=fast, adjust=False).mean()
    ema_slow   = series.ewm(span=slow, adjust=False).mean()
    macd_line  = ema_fast - ema_slow
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    histogram  = macd_line - signal_line
    return macd_line, signal_line, histogram


def compute_bollinger(series: pd.Series, period: int = 20, std: float = 2.0):
    mid  = series.rolling(period).mean()
    band = series.rolling(period).std()
    return mid, mid + std * band, mid - std * band


def compute_momentum(series: pd.Series, period: int = 10) -> pd.Series:
    return series.pct_change(period) * 100


def compute_atr(high, low, close, period: int = 14) -> pd.Series:
    tr = pd.concat([
        high - low,
        (high - close.shift()).abs(),
        (low  - close.shift()).abs(),
    ], axis=1).max(axis=1)
    return tr.rolling(period).mean()


def analyze_stock(ticker_sym: str, df: pd.DataFrame) -> dict:
    """Compute full technical analysis for one stock."""
    close  = df['Close']
    high   = df['High']
    low    = df['Low']
    volume = df['Volume']

    # Moving averages
    ma20  = close.rolling(20).mean()
    ma50  = close.rolling(50).mean()
    ema12 = close.ewm(span=12, adjust=False).mean()

    # RSI
    rsi = compute_rsi(close)

    # MACD
    macd_line, signal_line, histogram = compute_macd(close)

    # Bollinger Bands
    bb_mid, bb_upper, bb_lower = compute_bollinger(close)

    # Momentum
    momentum = compute_momentum(close, 10)

    # ATR (volatility)
    atr = compute_atr(high, low, close)

    # Volume MA
    vol_ma20 = volume.rolling(20).mean()

    # 52-week high/low (use all available data – max 1 year)
    w52_high = high.max()
    w52_low  = low.min()
    curr     = close.iloc[-1]
    pct_from_high = (curr - w52_high) / w52_high * 100
    pct_from_low  = (curr - w52_low)  / w52_low  * 100

    # Period return
    period_return = (curr - close.iloc[0]) / close.iloc[0] * 100

    # Trend signal
    trend = "BULLISH" if (curr > ma20.iloc[-1] > ma50.iloc[-1]) else \
            "BEARISH" if (curr < ma20.iloc[-1] < ma50.iloc[-1]) else "NEUTRAL"

    # RSI signal
    rsi_val = rsi.iloc[-1]
    rsi_sig = "OVERBOUGHT" if rsi_val > 70 else "OVERSOLD" if rsi_val < 30 else "NEUTRAL"

    # MACD signal
    macd_sig = "BUY"  if (macd_line.iloc[-1] > signal_line.iloc[-1] and
                           histogram.iloc[-1] > histogram.iloc[-2]) else \
               "SELL" if (macd_line.iloc[-1] < signal_line.iloc[-1]) else "NEUTRAL"

    # Momentum
    mom_val = momentum.iloc[-1]

    # Volume trend (last 5 days vs 20-day avg)
    vol_ratio = volume.iloc[-5:].mean() / vol_ma20.iloc[-1] if vol_ma20.iloc[-1] > 0 else 1.0

    # BB position
    bb_pct = (curr - bb_lower.iloc[-1]) / (bb_upper.iloc[-1] - bb_lower.iloc[-1]) \
             if (bb_upper.iloc[-1] - bb_lower.iloc[-1]) > 0 else 0.5

    return {
        "close":          close,
        "high":           high,
        "low":            low,
        "volume":         volume,
        "ma20":           ma20,
        "ma50":           ma50,
        "ema12":          ema12,
        "rsi":            rsi,
        "macd_line":      macd_line,
        "signal_line":    signal_line,
        "histogram":      histogram,
        "bb_mid":         bb_mid,
        "bb_upper":       bb_upper,
        "bb_lower":       bb_lower,
        "momentum":       momentum,
        "atr":            atr,
        "vol_ma20":       vol_ma20,
        # Scalars
        "current":        curr,
        "period_return":  period_return,
        "w52_high":       w52_high,
        "w52_low":        w52_low,
        "pct_from_high":  pct_from_high,
        "pct_from_low":   pct_from_low,
        "rsi_val":        rsi_val,
        "rsi_sig":        rsi_sig,
        "macd_sig":       macd_sig,
        "trend":          trend,
        "momentum_val":   mom_val,
        "vol_ratio":      vol_ratio,
        "bb_pct":         bb_pct,
        "atr_val":        atr.iloc[-1],
    }


# ─────────────────────────────────────────────────────────────────────────────
# CHART STYLING
# ─────────────────────────────────────────────────────────────────────────────
def apply_dark_style():
    plt.rcParams.update({
        "figure.facecolor":   PALETTE["bg"],
        "axes.facecolor":     PALETTE["surface"],
        "axes.edgecolor":     PALETTE["border"],
        "axes.labelcolor":    PALETTE["muted"],
        "axes.titlecolor":    PALETTE["text"],
        "xtick.color":        PALETTE["muted"],
        "ytick.color":        PALETTE["muted"],
        "grid.color":         PALETTE["border"],
        "grid.linewidth":     0.5,
        "grid.alpha":         0.6,
        "text.color":         PALETTE["text"],
        "font.family":        "monospace",
        "figure.autolayout":  False,
    })


def fmt_inr(val: float) -> str:
    if val >= 1e7:
        return f"₹{val/1e7:.1f}Cr"
    if val >= 1e5:
        return f"₹{val/1e5:.1f}L"
    return f"₹{val:.0f}"


def signal_color(sig: str) -> str:
    mapping = {
        "BULLISH": PALETTE["green"],  "BEARISH": PALETTE["red"],
        "NEUTRAL": PALETTE["amber"],
        "OVERBOUGHT": PALETTE["red"], "OVERSOLD": PALETTE["green"],
        "BUY":    PALETTE["green"],   "SELL": PALETTE["red"],
    }
    return mapping.get(sig, PALETTE["muted"])


# ─────────────────────────────────────────────────────────────────────────────
# CHART 1 – OVERVIEW DASHBOARD (all 50 stocks)
# ─────────────────────────────────────────────────────────────────────────────
def plot_overview(summary_df: pd.DataFrame, out_dir: str):
    apply_dark_style()
    fig = plt.figure(figsize=(24, 18))
    fig.patch.set_facecolor(PALETTE["bg"])

    gs = gridspec.GridSpec(3, 3, figure=fig,
                           hspace=0.45, wspace=0.35,
                           left=0.05, right=0.97, top=0.92, bottom=0.05)

    # ── Title ────────────────────────────────────────────────────────────────
    fig.text(0.5, 0.96, "NIFTY 50 — 3-Month Performance Dashboard",
             ha='center', va='top', fontsize=20, fontweight='bold',
             color=PALETTE["text"], fontfamily='monospace')
    fig.text(0.5, 0.935, f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M IST')}",
             ha='center', va='top', fontsize=11, color=PALETTE["muted"])

    # ── 1. Returns bar chart ─────────────────────────────────────────────────
    ax1 = fig.add_subplot(gs[0, :2])
    df_sorted = summary_df.sort_values('period_return')
    colors = [PALETTE["green"] if r >= 0 else PALETTE["red"]
              for r in df_sorted['period_return']]
    bars = ax1.barh(df_sorted['symbol'], df_sorted['period_return'],
                    color=colors, height=0.7, edgecolor='none')
    ax1.axvline(0, color=PALETTE["border"], linewidth=1)
    ax1.set_title("3-Month Price Returns (%)", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax1.set_xlabel("Return (%)", color=PALETTE["muted"])
    ax1.tick_params(axis='y', labelsize=7)
    ax1.tick_params(axis='x', labelsize=8)
    ax1.grid(axis='x', alpha=0.3)
    ax1.set_facecolor(PALETTE["surface"])

    # ── 2. Sector heatmap ────────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[0, 2])
    sector_ret = summary_df.groupby('sector')['period_return'].mean().sort_values()
    s_colors = [PALETTE["green"] if r >= 0 else PALETTE["red"]
                for r in sector_ret]
    ax2.barh(sector_ret.index, sector_ret.values, color=s_colors,
             height=0.65, edgecolor='none')
    ax2.axvline(0, color=PALETTE["border"], linewidth=1)
    ax2.set_title("Avg Return by Sector (%)", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax2.tick_params(labelsize=8)
    ax2.grid(axis='x', alpha=0.3)
    ax2.set_facecolor(PALETTE["surface"])

    # ── 3. RSI Distribution ──────────────────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 0])
    rsi_vals = summary_df['rsi_val'].dropna()
    ax3.hist(rsi_vals, bins=20, color=PALETTE["accent"], edgecolor=PALETTE["border"],
             linewidth=0.5, alpha=0.85)
    ax3.axvline(70, color=PALETTE["red"],   linestyle='--', linewidth=1.5, label='Overbought')
    ax3.axvline(30, color=PALETTE["green"], linestyle='--', linewidth=1.5, label='Oversold')
    ax3.axvline(50, color=PALETTE["muted"], linestyle=':',  linewidth=1,   label='Mid')
    ax3.set_title("RSI Distribution", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax3.set_xlabel("RSI (14)", color=PALETTE["muted"])
    ax3.set_ylabel("Count",    color=PALETTE["muted"])
    ax3.legend(fontsize=7, facecolor=PALETTE["card"], edgecolor=PALETTE["border"],
               labelcolor=PALETTE["text"])
    ax3.set_facecolor(PALETTE["surface"])

    # ── 4. RSI Signal Pie ────────────────────────────────────────────────────
    ax4 = fig.add_subplot(gs[1, 1])
    sig_counts = summary_df['rsi_sig'].value_counts()
    pie_colors = [signal_color(s) for s in sig_counts.index]
    wedges, texts, autotexts = ax4.pie(
        sig_counts.values, labels=sig_counts.index,
        colors=pie_colors, autopct='%1.0f%%',
        startangle=90, pctdistance=0.75,
        textprops={'color': PALETTE["text"], 'fontsize': 9},
        wedgeprops={'edgecolor': PALETTE["bg"], 'linewidth': 2}
    )
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color(PALETTE["bg"])
        at.set_fontweight('bold')
    ax4.set_title("RSI Signal Distribution", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax4.set_facecolor(PALETTE["bg"])

    # ── 5. Trend signal breakdown ────────────────────────────────────────────
    ax5 = fig.add_subplot(gs[1, 2])
    trend_counts = summary_df['trend'].value_counts()
    t_colors = [signal_color(t) for t in trend_counts.index]
    bars5 = ax5.bar(trend_counts.index, trend_counts.values,
                    color=t_colors, edgecolor=PALETTE["border"], linewidth=0.5)
    for bar, val in zip(bars5, trend_counts.values):
        ax5.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                 str(val), ha='center', va='bottom',
                 color=PALETTE["text"], fontsize=11, fontweight='bold')
    ax5.set_title("Price Trend Breakdown\n(MA20 vs MA50 vs Price)", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax5.set_ylabel("# Stocks", color=PALETTE["muted"])
    ax5.set_facecolor(PALETTE["surface"])

    # ── 6. 52-week position scatter ──────────────────────────────────────────
    ax6 = fig.add_subplot(gs[2, :2])
    x_pos = summary_df['pct_from_low']
    y_pos = summary_df['period_return']
    sc_colors = [signal_color(t) for t in summary_df['trend']]
    sc = ax6.scatter(x_pos, y_pos, c=sc_colors, s=70, alpha=0.85,
                     edgecolors=PALETTE["border"], linewidths=0.5, zorder=3)
    for _, row in summary_df.iterrows():
        ax6.annotate(row['symbol'],
                     (row['pct_from_low'], row['period_return']),
                     textcoords='offset points', xytext=(4, 2),
                     fontsize=5.5, color=PALETTE["muted"])
    ax6.axhline(0, color=PALETTE["border"], linewidth=0.8, linestyle='--')
    ax6.set_title("52-Week Low Recovery vs 3-Month Return  (colour = trend)",
                  fontsize=12, pad=8, color=PALETTE["text"], fontweight='bold')
    ax6.set_xlabel("% Above 52-Week Low", color=PALETTE["muted"])
    ax6.set_ylabel("3-Month Return (%)", color=PALETTE["muted"])
    ax6.grid(alpha=0.25)
    ax6.set_facecolor(PALETTE["surface"])
    handles = [mpatches.Patch(color=signal_color(t), label=t)
               for t in ["BULLISH", "NEUTRAL", "BEARISH"]]
    ax6.legend(handles=handles, fontsize=8, facecolor=PALETTE["card"],
               edgecolor=PALETTE["border"], labelcolor=PALETTE["text"])

    # ── 7. MACD signal pie ───────────────────────────────────────────────────
    ax7 = fig.add_subplot(gs[2, 2])
    macd_counts = summary_df['macd_sig'].value_counts()
    m_colors = [signal_color(s) for s in macd_counts.index]
    wedges7, texts7, autotexts7 = ax7.pie(
        macd_counts.values, labels=macd_counts.index,
        colors=m_colors, autopct='%1.0f%%',
        startangle=90, pctdistance=0.75,
        textprops={'color': PALETTE["text"], 'fontsize': 9},
        wedgeprops={'edgecolor': PALETTE["bg"], 'linewidth': 2}
    )
    for at in autotexts7:
        at.set_fontsize(8)
        at.set_color(PALETTE["bg"])
        at.set_fontweight('bold')
    ax7.set_title("MACD Signal Distribution", fontsize=12, pad=8,
                  color=PALETTE["text"], fontweight='bold')
    ax7.set_facecolor(PALETTE["bg"])

    plt.savefig(os.path.join(out_dir, "01_nifty50_overview.png"),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print("  ✓ Chart 01: Overview dashboard saved")


# ─────────────────────────────────────────────────────────────────────────────
# CHART 2 – TOP 10 GAINERS & LOSERS
# ─────────────────────────────────────────────────────────────────────────────
def plot_gainers_losers(summary_df: pd.DataFrame, out_dir: str):
    apply_dark_style()
    fig, axes = plt.subplots(1, 2, figsize=(18, 8))
    fig.patch.set_facecolor(PALETTE["bg"])
    fig.suptitle("NIFTY 50 — Top 10 Gainers & Losers (3-Month)",
                 fontsize=16, fontweight='bold', color=PALETTE["text"], y=0.98)

    df_sorted = summary_df.sort_values('period_return')
    losers  = df_sorted.head(10)
    gainers = df_sorted.tail(10).iloc[::-1]

    for ax, df_sub, title, clr in [
        (axes[0], losers,  "🔴  Biggest Losers", PALETTE["red"]),
        (axes[1], gainers, "🟢  Biggest Gainers", PALETTE["green"]),
    ]:
        bars = ax.barh(
            [f"{r['symbol']}\n{r['name'][:22]}" for _, r in df_sub.iterrows()],
            df_sub['period_return'], color=clr,
            height=0.65, edgecolor='none', alpha=0.9
        )
        # Value labels on bars
        for bar, val in zip(bars, df_sub['period_return']):
            xpos = bar.get_width()
            ha   = 'left' if val > 0 else 'right'
            offset = 0.3 if val > 0 else -0.3
            ax.text(xpos + offset, bar.get_y() + bar.get_height()/2,
                    f"{val:+.1f}%", va='center', ha=ha,
                    color=PALETTE["text"], fontsize=9, fontweight='bold')
        ax.axvline(0, color=PALETTE["muted"], linewidth=0.8)
        ax.set_title(title, fontsize=13, pad=10, color=PALETTE["text"])
        ax.tick_params(axis='y', labelsize=8)
        ax.tick_params(axis='x', labelsize=9)
        ax.grid(axis='x', alpha=0.25)
        ax.set_facecolor(PALETTE["surface"])
        # RSI annotation
        for i, (_, row) in enumerate(df_sub.iterrows()):
            ax.text(
                ax.get_xlim()[0] if clr == PALETTE["red"] else ax.get_xlim()[1],
                i,
                f"RSI:{row['rsi_val']:.0f}",
                va='center',
                ha='left' if clr == PALETTE["green"] else 'right',
                fontsize=7, color=PALETTE["muted"]
            )

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig(os.path.join(out_dir, "02_gainers_losers.png"),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print("  ✓ Chart 02: Gainers & Losers saved")


# ─────────────────────────────────────────────────────────────────────────────
# CHART 3 – TECHNICAL HEATMAP (RSI + Momentum + BB Position)
# ─────────────────────────────────────────────────────────────────────────────
def plot_technical_heatmap(summary_df: pd.DataFrame, out_dir: str):
    apply_dark_style()
    fig, axes = plt.subplots(1, 3, figsize=(22, 14))
    fig.patch.set_facecolor(PALETTE["bg"])
    fig.suptitle("NIFTY 50 — Technical Indicator Heatmap",
                 fontsize=16, fontweight='bold', color=PALETTE["text"], y=0.99)

    df_s = summary_df.sort_values('sector').reset_index(drop=True)
    metrics = [
        ("RSI (14)",           "rsi_val",      0,   100,  "RdYlGn"),
        ("10-Day Momentum (%)", "momentum_val", None, None, "RdYlGn"),
        ("BB Position (0–1)",  "bb_pct",       0,   1,    "coolwarm"),
    ]

    for ax, (title, col, vmin, vmax, cmap) in zip(axes, metrics):
        vals = df_s[col].values.reshape(-1, 1)
        im = ax.imshow(vals, aspect='auto', cmap=cmap,
                       vmin=vmin or vals.min(), vmax=vmax or vals.max())

        ax.set_yticks(range(len(df_s)))
        ax.set_yticklabels(
            [f"{r['symbol']:<12} [{r['sector'][:8]}]" for _, r in df_s.iterrows()],
            fontsize=7, fontfamily='monospace'
        )
        ax.set_xticks([])
        ax.set_title(title, fontsize=11, pad=8,
                     color=PALETTE["text"], fontweight='bold')

        # Value annotations
        for i, v in enumerate(df_s[col]):
            ax.text(0, i, f"{v:.1f}", ha='center', va='center',
                    fontsize=6.5, fontweight='bold',
                    color='white' if (abs(v - (50 if vmax == 100 else 0.5 if vmax == 1 else 0)) > 10) else 'black')

        cbar = fig.colorbar(im, ax=ax, shrink=0.4, pad=0.02)
        cbar.ax.tick_params(labelsize=7, colors=PALETTE["muted"])

    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.savefig(os.path.join(out_dir, "03_technical_heatmap.png"),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print("  ✓ Chart 03: Technical heatmap saved")


# ─────────────────────────────────────────────────────────────────────────────
# CHART 4 – INDIVIDUAL STOCK DEEP-DIVE (top gainer, top loser, user pick)
# ─────────────────────────────────────────────────────────────────────────────
def plot_stock_deepdive(symbol: str, name: str, analysis: dict, out_dir: str,
                        label: str = ""):
    apply_dark_style()
    fig = plt.figure(figsize=(20, 14))
    fig.patch.set_facecolor(PALETTE["bg"])

    gs = gridspec.GridSpec(4, 1, figure=fig, height_ratios=[3, 1, 1, 1],
                           hspace=0.08, left=0.07, right=0.97,
                           top=0.93, bottom=0.05)

    dates = analysis["close"].index
    c = analysis

    # ── Panel 1: Price + MAs + Bollinger ─────────────────────────────────────
    ax1 = fig.add_subplot(gs[0])
    ax1.fill_between(dates, c["bb_lower"], c["bb_upper"],
                     alpha=0.12, color=PALETTE["accent"], label='BB Bands')
    ax1.plot(dates, c["bb_upper"], color=PALETTE["accent"],
             linewidth=0.6, linestyle='--', alpha=0.6)
    ax1.plot(dates, c["bb_lower"], color=PALETTE["accent"],
             linewidth=0.6, linestyle='--', alpha=0.6)
    ax1.plot(dates, c["bb_mid"],   color=PALETTE["accent"],
             linewidth=0.8, linestyle=':', alpha=0.5, label='BB Mid')
    ax1.plot(dates, c["close"],  color=PALETTE["text"],
             linewidth=1.8, label='Close', zorder=5)
    ax1.plot(dates, c["ma20"],   color=PALETTE["amber"],
             linewidth=1.2, label='MA20', alpha=0.9)
    ax1.plot(dates, c["ma50"],   color=PALETTE["purple"],
             linewidth=1.2, label='MA50', alpha=0.9)
    ax1.plot(dates, c["ema12"],  color=PALETTE["teal"],
             linewidth=1.0, label='EMA12', alpha=0.7, linestyle='-.')

    # 52W markers
    w52_date_h = c["high"].idxmax()
    w52_date_l = c["low"].idxmin()
    ax1.annotate(f"52W High\n{c['w52_high']:.1f}",
                 xy=(w52_date_h, c["w52_high"]),
                 xytext=(10, 10), textcoords='offset points',
                 color=PALETTE["green"], fontsize=7,
                 arrowprops=dict(arrowstyle='->', color=PALETTE["green"], lw=0.8))
    ax1.annotate(f"52W Low\n{c['w52_low']:.1f}",
                 xy=(w52_date_l, c["w52_low"]),
                 xytext=(10, -20), textcoords='offset points',
                 color=PALETTE["red"], fontsize=7,
                 arrowprops=dict(arrowstyle='->', color=PALETTE["red"], lw=0.8))

    ret_color = PALETTE["green"] if c["period_return"] >= 0 else PALETTE["red"]
    ax1.set_title(
        f"{label}  {symbol} — {name}  |  "
        f"Current: ₹{c['current']:.1f}  |  "
        f"3M Return: {c['period_return']:+.1f}%  |  "
        f"Trend: {c['trend']}",
        fontsize=12, pad=8, color=PALETTE["text"], fontweight='bold'
    )
    ax1.legend(loc='upper left', fontsize=8, facecolor=PALETTE["card"],
               edgecolor=PALETTE["border"], labelcolor=PALETTE["text"],
               ncol=4)
    ax1.set_ylabel("Price (₹)", color=PALETTE["muted"])
    ax1.tick_params(labelbottom=False)
    ax1.grid(alpha=0.2)
    ax1.set_facecolor(PALETTE["surface"])

    # ── Panel 2: Volume ───────────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[1], sharex=ax1)
    vol_colors = [PALETTE["green"] if c["close"].iloc[i] >= c["close"].iloc[i-1]
                  else PALETTE["red"]
                  for i in range(len(c["close"]))]
    ax2.bar(dates, c["volume"], color=vol_colors, alpha=0.7, width=0.8)
    ax2.plot(dates, c["vol_ma20"], color=PALETTE["amber"],
             linewidth=1.2, label='Vol MA20')
    ax2.set_ylabel("Volume", color=PALETTE["muted"])
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda x, _: f"{x/1e6:.0f}M" if x >= 1e6 else f"{x/1e3:.0f}K"))
    ax2.tick_params(labelbottom=False)
    ax2.grid(alpha=0.2)
    ax2.set_facecolor(PALETTE["surface"])
    ax2.legend(fontsize=7, facecolor=PALETTE["card"],
               edgecolor=PALETTE["border"], labelcolor=PALETTE["text"])

    # ── Panel 3: RSI ──────────────────────────────────────────────────────────
    ax3 = fig.add_subplot(gs[2], sharex=ax1)
    ax3.plot(dates, c["rsi"], color=PALETTE["accent"], linewidth=1.5, label='RSI(14)')
    ax3.axhline(70, color=PALETTE["red"],   linestyle='--', linewidth=1, alpha=0.8)
    ax3.axhline(30, color=PALETTE["green"], linestyle='--', linewidth=1, alpha=0.8)
    ax3.axhline(50, color=PALETTE["muted"], linestyle=':',  linewidth=0.7)
    ax3.fill_between(dates, c["rsi"], 70,
                     where=c["rsi"] > 70, alpha=0.25, color=PALETTE["red"])
    ax3.fill_between(dates, c["rsi"], 30,
                     where=c["rsi"] < 30, alpha=0.25, color=PALETTE["green"])
    ax3.set_ylim(0, 100)
    ax3.set_ylabel("RSI", color=PALETTE["muted"])
    ax3.text(dates[-1], 72, f"OB(70)", fontsize=7, color=PALETTE["red"],   ha='right')
    ax3.text(dates[-1], 32, f"OS(30)", fontsize=7, color=PALETTE["green"], ha='right')
    current_rsi = c["rsi"].iloc[-1]
    ax3.text(dates[-1], current_rsi,
             f" {current_rsi:.1f} [{c['rsi_sig']}]",
             fontsize=8, color=signal_color(c["rsi_sig"]), va='center')
    ax3.tick_params(labelbottom=False)
    ax3.grid(alpha=0.2)
    ax3.set_facecolor(PALETTE["surface"])

    # ── Panel 4: MACD ─────────────────────────────────────────────────────────
    ax4 = fig.add_subplot(gs[3], sharex=ax1)
    ax4.plot(dates, c["macd_line"],   color=PALETTE["accent"], linewidth=1.4, label='MACD')
    ax4.plot(dates, c["signal_line"], color=PALETTE["orange"], linewidth=1.2, label='Signal')
    hist_colors = [PALETTE["green"] if h >= 0 else PALETTE["red"]
                   for h in c["histogram"]]
    ax4.bar(dates, c["histogram"], color=hist_colors, alpha=0.6,
            width=0.8, label='Histogram')
    ax4.axhline(0, color=PALETTE["border"], linewidth=0.8)
    ax4.set_ylabel("MACD", color=PALETTE["muted"])
    ax4.set_xlabel("Date", color=PALETTE["muted"])
    ax4.legend(loc='upper left', fontsize=7, facecolor=PALETTE["card"],
               edgecolor=PALETTE["border"], labelcolor=PALETTE["text"])
    ax4.grid(alpha=0.2)
    ax4.set_facecolor(PALETTE["surface"])

    # Format x-axis dates
    ax4.xaxis.set_major_formatter(matplotlib.dates.DateFormatter('%d %b'))
    ax4.xaxis.set_major_locator(matplotlib.dates.WeekdayLocator(interval=2))
    plt.setp(ax4.xaxis.get_majorticklabels(), rotation=30, ha='right', fontsize=7)

    fname = f"04_deepdive_{symbol.replace('&','').replace('-','_')}.png"
    plt.savefig(os.path.join(out_dir, fname),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print(f"  ✓ Chart 04: Deep-dive [{symbol}] saved")


# ─────────────────────────────────────────────────────────────────────────────
# CHART 5 – 52-WEEK HIGH/LOW TRACKER
# ─────────────────────────────────────────────────────────────────────────────
def plot_52week(summary_df: pd.DataFrame, out_dir: str):
    apply_dark_style()
    fig, ax = plt.subplots(figsize=(20, 14))
    fig.patch.set_facecolor(PALETTE["bg"])

    df_s = summary_df.sort_values('pct_from_high').reset_index(drop=True)
    y = range(len(df_s))

    # Range bar (52W low to 52W high)
    for i, row in df_s.iterrows():
        norm_low  = 0
        norm_curr = row['pct_from_low']
        norm_high = row['pct_from_low'] + abs(row['pct_from_high'])
        ax.barh(i, norm_high, left=norm_low, height=0.5,
                color=PALETTE["border"], alpha=0.4)
        ax.scatter(norm_curr, i, color=PALETTE["accent"], s=50, zorder=5)

    ax.set_yticks(list(y))
    ax.set_yticklabels(
        [f"{r['symbol']:<12} ({r['sector'][:6]})" for _, r in df_s.iterrows()],
        fontsize=7.5, fontfamily='monospace'
    )
    ax.set_title("52-Week Price Range Tracker\n"
                 "Bar = 52W Low → 52W High  |  Dot = Current Price",
                 fontsize=14, pad=10, color=PALETTE["text"], fontweight='bold')
    ax.set_xlabel("% Above 52-Week Low", color=PALETTE["muted"])
    ax.grid(axis='x', alpha=0.2)
    ax.set_facecolor(PALETTE["surface"])

    # Annotations for near-high / near-low
    for i, row in df_s.iterrows():
        if row['pct_from_high'] > -5:
            ax.text(row['pct_from_low'] + 0.5, i, "NEAR HIGH",
                    va='center', fontsize=6, color=PALETTE["green"], fontweight='bold')
        if row['pct_from_high'] < -30:
            ax.text(row['pct_from_low'] + 0.5, i, "DEEP CORRECTION",
                    va='center', fontsize=6, color=PALETTE["red"])

    legend_elems = [
        mpatches.Patch(color=PALETTE["border"], alpha=0.4, label='52W Range'),
        plt.Line2D([0], [0], marker='o', color='w',
                   markerfacecolor=PALETTE["accent"], markersize=7, label='Current'),
    ]
    ax.legend(handles=legend_elems, fontsize=9, facecolor=PALETTE["card"],
              edgecolor=PALETTE["border"], labelcolor=PALETTE["text"])

    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, "05_52week_tracker.png"),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print("  ✓ Chart 05: 52-Week tracker saved")


# ─────────────────────────────────────────────────────────────────────────────
# CHART 6 – SECTOR MOMENTUM GRID
# ─────────────────────────────────────────────────────────────────────────────
def plot_sector_grid(summary_df: pd.DataFrame, out_dir: str):
    apply_dark_style()
    sectors = summary_df['sector'].unique()
    n_cols  = 4
    n_rows  = int(np.ceil(len(sectors) / n_cols))
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(22, n_rows * 4))
    fig.patch.set_facecolor(PALETTE["bg"])
    fig.suptitle("NIFTY 50 — Sector-wise Stock Returns (3-Month)",
                 fontsize=16, fontweight='bold', color=PALETTE["text"], y=0.99)

    axes_flat = axes.flatten() if n_rows > 1 else axes
    for ax in axes_flat:
        ax.set_facecolor(PALETTE["surface"])

    for idx, sector in enumerate(sorted(sectors)):
        ax  = axes_flat[idx]
        sub = summary_df[summary_df['sector'] == sector].sort_values('period_return')
        clrs = [PALETTE["green"] if r >= 0 else PALETTE["red"]
                for r in sub['period_return']]
        bars = ax.bar(sub['symbol'], sub['period_return'],
                      color=clrs, edgecolor='none', alpha=0.9)
        ax.axhline(0, color=PALETTE["border"], linewidth=0.8)
        ax.set_title(f"{'●'} {sector}", fontsize=10, color=SECTOR_COLORS.get(sector, PALETTE["accent"]),
                     fontweight='bold', pad=6)
        ax.tick_params(axis='x', labelsize=7, rotation=30)
        ax.tick_params(axis='y', labelsize=7)
        ax.set_ylabel("Return %", color=PALETTE["muted"], fontsize=8)
        ax.grid(axis='y', alpha=0.2)
        for bar, val in zip(bars, sub['period_return']):
            ax.text(bar.get_x() + bar.get_width()/2,
                    bar.get_height() + (0.3 if val >= 0 else -1.2),
                    f"{val:+.1f}%", ha='center', va='bottom',
                    fontsize=6.5, color=PALETTE["text"])

    # Hide empty subplots
    for idx in range(len(sectors), len(axes_flat)):
        axes_flat[idx].set_visible(False)

    plt.tight_layout(rect=[0, 0, 1, 0.97])
    plt.savefig(os.path.join(out_dir, "06_sector_grid.png"),
                dpi=150, bbox_inches='tight', facecolor=PALETTE["bg"])
    plt.close()
    print("  ✓ Chart 06: Sector grid saved")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(summary_df: pd.DataFrame, analyses: dict,
                 out_dir: str, period: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.styles.differential import DifferentialStyle

    wb = Workbook()
    wb.remove(wb.active)

    DARK   = '0D1117'
    MID    = '1C2128'
    BORDER = '30363D'
    BLUE   = '58A6FF'
    GREEN_ = '3FB950'
    RED_   = 'F85149'
    AMBER_ = 'D29922'
    WHITE_ = 'E6EDF3'
    GRAY_  = '8B949E'

    def hfont(size=10, bold=True, color=WHITE_):
        return Font(name='Consolas', size=size, bold=bold, color=color)
    def dfont(size=9, bold=False, color=WHITE_):
        return Font(name='Consolas', size=size, bold=bold, color=color)
    def hfill(color=DARK):
        return PatternFill('solid', fgColor=color)
    def center():
        return Alignment(horizontal='center', vertical='center')
    def left():
        return Alignment(horizontal='left', vertical='center')
    def tb():
        s = Side(style='thin', color=BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet('Summary')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    headers = [
        'Symbol', 'Company Name', 'Sector',
        'Current Price (₹)', '3M Return (%)', 'RSI (14)', 'RSI Signal',
        'MACD Signal', 'Trend', 'Momentum (%)', 'BB Position',
        '52W High (₹)', '52W Low (₹)', '% from 52W High', '% from 52W Low',
        'ATR (₹)', 'Vol Ratio (5d/20d)',
    ]

    # Title
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']
    t.value = f'NIFTY 50 — Technical Analysis Summary  |  Period: {period.upper()}  |  Generated: {datetime.now().strftime("%d %b %Y")}'
    t.font = hfont(size=12)
    t.fill = hfill(DARK)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font  = hfont(size=9)
        c.fill  = hfill('161B22')
        c.alignment = center()
        c.border = tb()
    ws.row_dimensions[2].height = 24

    # Data
    for ri, row in summary_df.sort_values('sector').reset_index(drop=True).iterrows():
        dr = ri + 3
        bg = MID if ri % 2 == 0 else DARK

        def w(col, val, fmt=None, num=False, bold=False, color=WHITE_):
            c2 = ws.cell(row=dr, column=col, value=val)
            c2.font = dfont(bold=bold, color=color)
            c2.fill = hfill(bg)
            c2.alignment = center() if col > 1 else left()
            c2.border = tb()
            if fmt and num:
                c2.number_format = fmt
            return c2

        ret_clr  = GREEN_ if row['period_return'] >= 0 else RED_
        rsi_clr  = RED_ if row['rsi_val'] > 70 else (GREEN_ if row['rsi_val'] < 30 else AMBER_)
        trd_clr  = (GREEN_ if row['trend'] == 'BULLISH' else
                    RED_   if row['trend'] == 'BEARISH' else AMBER_)

        w(1,  row['symbol'],        bold=True, color=BLUE)
        w(2,  row['name'],          color=WHITE_)
        w(3,  row['sector'],        color=GRAY_)
        w(4,  row['current'],       fmt='#,##0.00', num=True)
        w(5,  row['period_return'], fmt='0.0%',     num=True,  bold=True, color=ret_clr)
        w(6,  row['rsi_val'],       fmt='0.0',      num=True,  color=rsi_clr)
        w(7,  row['rsi_sig'],       color=rsi_clr)
        w(8,  row['macd_sig'],
            color=(GREEN_ if row['macd_sig']=='BUY' else RED_ if row['macd_sig']=='SELL' else AMBER_))
        w(9,  row['trend'],         color=trd_clr)
        w(10, row['momentum_val'],  fmt='0.0',      num=True,
            color=(GREEN_ if row['momentum_val'] >= 0 else RED_))
        w(11, row['bb_pct'],        fmt='0.00',     num=True)
        w(12, row['w52_high'],      fmt='#,##0.00', num=True)
        w(13, row['w52_low'],       fmt='#,##0.00', num=True)
        w(14, row['pct_from_high'], fmt='0.0%',     num=True,
            color=(GREEN_ if row['pct_from_high'] > -5 else RED_))
        w(15, row['pct_from_low'],  fmt='0.0%',     num=True)
        w(16, row['atr_val'],       fmt='#,##0.00', num=True)
        w(17, row['vol_ratio'],     fmt='0.00x',    num=True)
        ws.row_dimensions[dr].height = 18

    # Column widths
    col_widths = [12,28,16,14,12,10,12,12,10,12,12,14,14,16,16,10,14]
    for ci, w_val in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w_val

    # ── Sheet 2: Price History ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Price History')
    ws2.sheet_view.showGridLines = False

    symbols = list(analyses.keys())
    # Write dates in column A
    ws2['A1'].value = 'Date'
    ws2['A1'].font  = hfont()
    ws2['A1'].fill  = hfill(DARK)
    ws2['A1'].alignment = center()

    for ci, sym in enumerate(symbols, 2):
        c = ws2.cell(row=1, column=ci, value=sym)
        c.font = hfont(size=9)
        c.fill = hfill(DARK)
        c.alignment = center()

    dates = analyses[symbols[0]]["close"].index
    for ri, date in enumerate(dates, 2):
        ws2.cell(row=ri, column=1, value=date.strftime('%d-%b-%Y')).font = dfont()
        ws2.cell(row=ri, column=1).fill = hfill(MID if ri % 2 == 0 else DARK)

    for ci, sym in enumerate(symbols, 2):
        close_vals = analyses[sym]["close"].values
        for ri, val in enumerate(close_vals, 2):
            c = ws2.cell(row=ri, column=ci, value=round(float(val), 2))
            c.font   = dfont()
            c.fill   = hfill(MID if ri % 2 == 0 else DARK)
            c.number_format = '#,##0.00'

    ws2.column_dimensions['A'].width = 14
    for ci in range(2, len(symbols)+2):
        ws2.column_dimensions[get_column_letter(ci)].width = 10

    # ── Sheet 3: Technical Indicators ─────────────────────────────────────────
    ws3 = wb.create_sheet('Technical Data')
    ws3.sheet_view.showGridLines = False

    tech_headers = ['Date'] + [f"{sym}_RSI" for sym in symbols] + \
                   [f"{sym}_MACD" for sym in symbols]
    ws3.cell(row=1, column=1, value=f'Technical Indicators — {period.upper()}').font = hfont()
    ws3.cell(row=1, column=1).fill = hfill(DARK)
    ws3.merge_cells(f'A1:{get_column_letter(len(tech_headers))}1')
    ws3.cell(row=1, column=1).alignment = center()

    date_col = analyses[symbols[0]]["close"].index
    for ri, date in enumerate(date_col, 2):
        ws3.cell(row=ri, column=1, value=date.strftime('%d-%b-%Y'))

    for ci, sym in enumerate(symbols, 2):
        ws3.cell(row=1, column=ci, value=f"{sym} RSI(14)")
        for ri, val in enumerate(analyses[sym]["rsi"].values, 2):
            c = ws3.cell(row=ri, column=ci, value=round(float(val), 1) if not np.isnan(val) else None)
            c.number_format = '0.0'

    macd_start = len(symbols) + 2
    for ci, sym in enumerate(symbols, macd_start):
        ws3.cell(row=1, column=ci, value=f"{sym} MACD")
        for ri, val in enumerate(analyses[sym]["macd_line"].values, 2):
            c = ws3.cell(row=ri, column=ci, value=round(float(val), 3) if not np.isnan(val) else None)
            c.number_format = '0.000'

    # ── Sheet 4: Sector Analysis ──────────────────────────────────────────────
    ws4 = wb.create_sheet('Sector Analysis')
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells('A1:F1')
    ws4['A1'].value = 'Sector-wise Performance Summary'
    ws4['A1'].font  = hfont(size=12)
    ws4['A1'].fill  = hfill(DARK)
    ws4['A1'].alignment = center()
    ws4.row_dimensions[1].height = 26

    sec_headers = ['Sector', '# Stocks', 'Avg Return (%)',
                   'Best Performer', 'Worst Performer', 'Avg RSI']
    for ci, h in enumerate(sec_headers, 1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.font = hfont(size=9)
        c.fill = hfill('161B22')
        c.alignment = center()
        c.border = tb()
    ws4.row_dimensions[2].height = 22

    sector_data = summary_df.groupby('sector').agg(
        count=('symbol','count'),
        avg_ret=('period_return','mean'),
        avg_rsi=('rsi_val','mean'),
    ).reset_index()

    for ri, row in sector_data.sort_values('avg_ret', ascending=False).iterrows():
        dr   = ri + 3
        bg   = MID if ri % 2 == 0 else DARK
        sec  = row['sector']
        sec_stocks = summary_df[summary_df['sector'] == sec]
        best  = sec_stocks.loc[sec_stocks['period_return'].idxmax(), 'symbol']
        worst = sec_stocks.loc[sec_stocks['period_return'].idxmin(), 'symbol']

        def sw(col, val, fmt=None, color=WHITE_):
            c2 = ws4.cell(row=dr, column=col, value=val)
            c2.font = dfont(color=color)
            c2.fill = hfill(bg)
            c2.alignment = center()
            c2.border = tb()
            if fmt: c2.number_format = fmt
        ret_clr = GREEN_ if row['avg_ret'] >= 0 else RED_
        sw(1, sec, color=SECTOR_COLORS.get(sec, WHITE_).lstrip('#'))
        sw(2, int(row['count']))
        sw(3, row['avg_ret'] / 100, fmt='0.0%', color=ret_clr)
        sw(4, best,  color=GREEN_)
        sw(5, worst, color=RED_)
        sw(6, round(row['avg_rsi'], 1))
        ws4.row_dimensions[dr].height = 20

    for ci, w in enumerate([18,10,14,16,16,10], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # Tab colors
    ws.sheet_properties.tabColor  = '58A6FF'
    ws2.sheet_properties.tabColor = '3FB950'
    ws3.sheet_properties.tabColor = 'D29922'
    ws4.sheet_properties.tabColor = 'BC8CFF'

    path = os.path.join(out_dir, 'nifty50_analysis.xlsx')
    wb.save(path)
    print(f"  ✓ Excel report saved → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Nifty 50 Technical Analyzer')
    parser.add_argument('--period', default='3mo',
                        choices=['1mo','3mo','6mo','1y'],
                        help='Data period (default: 3mo)')
    parser.add_argument('--stock',  default=None,
                        help='Extra single-stock deep-dive (e.g. RELIANCE)')
    parser.add_argument('--out',    default='nifty50_output',
                        help='Output folder (default: nifty50_output)')
    args = parser.parse_args()

    # Lazy import yfinance (so the script fails gracefully if missing)
    try:
        import yfinance as yf
    except ImportError:
        print("\n❌  yfinance not found.\n"
              "    Install with:  pip install yfinance\n")
        sys.exit(1)

    out_dir = args.out
    os.makedirs(out_dir, exist_ok=True)
    print(f"\n{'═'*60}")
    print(f"  NIFTY 50 TREND ANALYZER  |  Period: {args.period}")
    print(f"{'═'*60}\n")

    # ── Fetch data ─────────────────────────────────────────────────────────────
    print("📥  Fetching data from Yahoo Finance …")
    tickers_yf = [f"{sym}.NS" for sym in NIFTY50.keys()]
    raw = yf.download(
        tickers_yf,
        period=args.period,
        interval='1d',
        group_by='ticker',
        auto_adjust=True,
        threads=True,
        progress=True,
    )
    print()

    # ── Analyze each stock ─────────────────────────────────────────────────────
    print("⚙️   Computing technical indicators …\n")
    analyses = {}
    summary_rows = []

    for sym, (name, sector) in NIFTY50.items():
        ticker_key = f"{sym}.NS"
        try:
            if isinstance(raw.columns, pd.MultiIndex):
                df = raw[ticker_key].dropna(subset=['Close'])
            else:
                df = raw.dropna(subset=['Close'])

            if df.empty or len(df) < 20:
                print(f"  ⚠  {sym}: insufficient data, skipping")
                continue

            a = analyze_stock(sym, df)
            analyses[sym] = a

            summary_rows.append({
                'symbol':        sym,
                'name':          name,
                'sector':        sector,
                'current':       a['current'],
                'period_return': a['period_return'] / 100,  # store as decimal for Excel %
                'rsi_val':       a['rsi_val'],
                'rsi_sig':       a['rsi_sig'],
                'macd_sig':      a['macd_sig'],
                'trend':         a['trend'],
                'momentum_val':  a['momentum_val'],
                'bb_pct':        a['bb_pct'],
                'w52_high':      a['w52_high'],
                'w52_low':       a['w52_low'],
                'pct_from_high': a['pct_from_high'] / 100,
                'pct_from_low':  a['pct_from_low']  / 100,
                'atr_val':       a['atr_val'],
                'vol_ratio':     a['vol_ratio'],
            })
            print(f"  ✓ {sym:<14} {name[:28]:<28} "
                  f"Return: {a['period_return']:+6.1f}%  "
                  f"RSI: {a['rsi_val']:5.1f}  "
                  f"Trend: {a['trend']}")

        except Exception as e:
            print(f"  ⚠  {sym}: {e}")

    if not summary_rows:
        print("\n❌  No data retrieved. Check internet connection and try again.\n")
        sys.exit(1)

    summary_df = pd.DataFrame(summary_rows)
    # Raw percent columns for charts
    summary_df['period_return_pct'] = summary_df['period_return'] * 100
    summary_df['pct_from_high_pct'] = summary_df['pct_from_high'] * 100
    summary_df['pct_from_low_pct']  = summary_df['pct_from_low']  * 100

    # ── Terminal summary ───────────────────────────────────────────────────────
    print(f"\n{'─'*60}")
    print(f"  QUICK SUMMARY  |  {len(analyses)} stocks analyzed")
    print(f"{'─'*60}")
    bull = (summary_df['trend'] == 'BULLISH').sum()
    bear = (summary_df['trend'] == 'BEARISH').sum()
    neut = (summary_df['trend'] == 'NEUTRAL').sum()
    overbought = (summary_df['rsi_val'] > 70).sum()
    oversold   = (summary_df['rsi_val'] < 30).sum()
    avg_ret    = summary_df['period_return_pct'].mean()
    best_sym   = summary_df.loc[summary_df['period_return_pct'].idxmax()]
    worst_sym  = summary_df.loc[summary_df['period_return_pct'].idxmin()]

    print(f"  Market Breadth :  🟢 Bullish: {bull}  🟡 Neutral: {neut}  🔴 Bearish: {bear}")
    print(f"  RSI Extremes   :  Overbought: {overbought}  |  Oversold: {oversold}")
    print(f"  Avg 3M Return  :  {avg_ret:+.1f}%")
    print(f"  Top Gainer     :  {best_sym['symbol']} ({best_sym['period_return_pct']:+.1f}%)")
    print(f"  Top Loser      :  {worst_sym['symbol']} ({worst_sym['period_return_pct']:+.1f}%)")
    print(f"{'─'*60}\n")

    # ── Generate charts ────────────────────────────────────────────────────────
    print("🎨  Generating charts …\n")
    # Use raw pct for chart functions
    chart_df = summary_df.copy()
    chart_df['period_return'] = chart_df['period_return_pct']
    chart_df['pct_from_high'] = chart_df['pct_from_high_pct']
    chart_df['pct_from_low']  = chart_df['pct_from_low_pct']

    plot_overview(chart_df, out_dir)
    plot_gainers_losers(chart_df, out_dir)
    plot_technical_heatmap(chart_df, out_dir)

    # Deep-dive: top gainer
    best_sym_key  = summary_df.loc[summary_df['period_return_pct'].idxmax(), 'symbol']
    worst_sym_key = summary_df.loc[summary_df['period_return_pct'].idxmin(), 'symbol']
    plot_stock_deepdive(best_sym_key,
                        NIFTY50[best_sym_key][0],
                        analyses[best_sym_key], out_dir,
                        label="🏆 TOP GAINER")
    plot_stock_deepdive(worst_sym_key,
                        NIFTY50[worst_sym_key][0],
                        analyses[worst_sym_key], out_dir,
                        label="⬇ TOP LOSER")

    # User-specified deep-dive
    if args.stock and args.stock.upper() in analyses:
        s = args.stock.upper()
        plot_stock_deepdive(s, NIFTY50[s][0], analyses[s], out_dir,
                            label="🔍 SELECTED")

    plot_52week(chart_df, out_dir)
    plot_sector_grid(chart_df, out_dir)

    # ── Excel export ───────────────────────────────────────────────────────────
    print("\n📊  Exporting Excel report …\n")
    export_excel(summary_df, analyses, out_dir, args.period)

    print(f"\n{'═'*60}")
    print(f"  ✅  ALL DONE!  Output folder: ./{out_dir}/")
    print(f"{'═'*60}")
    print(f"  📁  Charts    : {out_dir}/01_nifty50_overview.png")
    print(f"               : {out_dir}/02_gainers_losers.png")
    print(f"               : {out_dir}/03_technical_heatmap.png")
    print(f"               : {out_dir}/04_deepdive_*.png")
    print(f"               : {out_dir}/05_52week_tracker.png")
    print(f"               : {out_dir}/06_sector_grid.png")
    print(f"  📊  Excel     : {out_dir}/nifty50_analysis.xlsx")
    print(f"{'═'*60}\n")


if __name__ == '__main__':
    main()
PYEOF
